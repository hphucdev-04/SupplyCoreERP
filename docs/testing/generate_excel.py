import os
import re
import json
import xml.etree.ElementTree as ET
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Đường dẫn thư mục làm việc
WORKSPACE_DIR = r"D:\ProjectOwner\SupplyCoreERP"
TEST_DIRS = [
    os.path.join(WORKSPACE_DIR, "test", "SupplyCoreERP.Domain.Tests"),
    os.path.join(WORKSPACE_DIR, "test", "SupplyCoreERP.EntityFrameworkCore.Tests"),
    os.path.join(WORKSPACE_DIR, "test", "SupplyCoreERP.Application.Tests")
]
MAPPING_FILE = os.path.join(WORKSPACE_DIR, "docs", "testing", "testcase_mapping.json")
EXCEL_OUTPUT = os.path.join(WORKSPACE_DIR, "docs", "testing", "Partner_and_Catalog_Tests_Spec.xlsx")
TRX_DIR = os.path.join(WORKSPACE_DIR, "docs", "testing", "test_result")

def parse_qa_test_attribute(attr_content):
    """
    Parse nội dung chuỗi bên trong [QATest(...)] để trích xuất các tham số:
    scenario, feature, layer, priority, note.
    Hỗ trợ cả named parameters (scenario: "...") và positional parameters.
    """
    # 1. Tìm các named parameters bằng Regex
    named_params = {}
    for param_name in ["scenario", "feature", "layer", "priority", "note"]:
        match = re.search(r'\b' + param_name + r'\s*:\s*"(.*?)"', attr_content, re.DOTALL)
        if match:
            named_params[param_name] = match.group(1).strip()
            
    # 2. Tìm tất cả các chuỗi ngoặc kép để xử lý positional parameters
    all_strings = re.findall(r'"(.*?)"', attr_content, re.DOTALL)
    
    # Gán giá trị mặc định
    scenario = named_params.get("scenario")
    feature = named_params.get("feature")
    layer = named_params.get("layer")
    priority = named_params.get("priority", "Medium")
    note = named_params.get("note", "Đã kiểm chứng thành công.")
    
    # Điền giá trị từ positional parameters nếu thiếu
    string_idx = 0
    if not scenario and string_idx < len(all_strings):
        scenario = all_strings[string_idx]
        string_idx += 1
    if not feature and string_idx < len(all_strings):
        feature = all_strings[string_idx]
        string_idx += 1
    if not layer and string_idx < len(all_strings):
        layer = all_strings[string_idx]
        string_idx += 1
    if priority == "Medium" and string_idx < len(all_strings):
        priority = all_strings[string_idx]
        string_idx += 1
    if note == "Đã kiểm chứng thành công." and string_idx < len(all_strings):
        note = all_strings[string_idx]
        string_idx += 1
        
    return {
        "scenario": scenario or "Mô tả kịch bản kiểm thử.",
        "feature": feature or "General",
        "layer": layer or "Domain",
        "priority": priority,
        "note": note
    }

def scan_cs_files():
    """
    Quét đệ quy các thư mục test để tìm các file C# (*.cs) chứa [QATest] 
    và trích xuất metadata kiểm thử.
    """
    test_cases = []
    
    # Regex tìm [QATest(...)] và phương thức kiểm thử ngay sau nó
    # re.DOTALL cho phép dấu chấm khớp với cả ký tự xuống dòng
    qa_test_pattern = re.compile(
        r'\[QATest\((.*?)\)\]\s*(?:\[Fact\]|\[Theory\])?\s*public\s+(?:async\s+)?(?:Task(?:\s*<.*?>)?|void)\s+([A-Za-z0-9_]+)\s*\(',
        re.DOTALL
    )
    
    # Dự phòng trường hợp [Fact] đứng trước [QATest]
    qa_test_pattern_alt = re.compile(
        r'(?:\[Fact\]|\[Theory\])\s*\[QATest\((.*?)\)\]\s*public\s+(?:async\s+)?(?:Task(?:\s*<.*?>)?|void)\s+([A-Za-z0-9_]+)\s*\(',
        re.DOTALL
    )

    for test_dir in TEST_DIRS:
        if not os.path.exists(test_dir):
            continue
            
        is_efcore = "EntityFrameworkCore.Tests" in test_dir
        is_app = "Application.Tests" in test_dir
        
        if is_efcore:
            test_type = "Integration Test"
        elif is_app:
            test_type = "API Test"
        else:
            test_type = "Unit Test"
        
        for root_path, _, files in os.walk(test_dir):
            for file in files:
                if not file.endswith(".cs") or file == "QATestAttribute.cs":
                    continue
                    
                file_path = os.path.join(root_path, file)
                with open(file_path, "r", encoding="utf-8") as f:
                    content = f.read()
                    
                # Quét mẫu 1
                for match in qa_test_pattern.finditer(content):
                    attr_content = match.group(1)
                    method_name = match.group(2)
                    meta = parse_qa_test_attribute(attr_content)
                    meta["method_name"] = method_name
                    meta["type"] = test_type
                    meta["class_name"] = os.path.splitext(file)[0]
                    meta["filepath"] = file_path
                    test_cases.append(meta)
                    
                # Quét mẫu 2 (dự phòng)
                for match in qa_test_pattern_alt.finditer(content):
                    attr_content = match.group(1)
                    method_name = match.group(2)
                    meta = parse_qa_test_attribute(attr_content)
                    if not any(tc["method_name"] == method_name for tc in test_cases):
                        meta["method_name"] = method_name
                        meta["type"] = test_type
                        meta["class_name"] = os.path.splitext(file)[0]
                        meta["filepath"] = file_path
                        test_cases.append(meta)
                        
    return test_cases

def get_group_prefix(tc):
    """
    Tính toán tiền tố nhóm ID dựa trên Feature, Layer và Loại test.
    """
    feature = tc["feature"]
    layer = tc["layer"]
    test_type = tc["type"]
    class_name = tc["class_name"]
    
    # Định nghĩa bảng viết tắt Feature
    feature_map = {
        "Medicine": "MED",
        "Supplier": "SUP",
        "Customer": "CUS",
        "SupplierProduct": "SP",
        "SupplierProductCondition": "SPC",
        "MedicineAppService": "MEDAPI"
    }
    
    # Xác định context (Catalog hay Partner)
    is_partner = "Partner" in tc["filepath"] or feature in ["Supplier", "Customer", "SupplierProduct", "SupplierProductCondition"]
    context_prefix = "PART" if is_partner else "CAT"
    
    # Xác định Type (UT, IT hoặc API)
    if test_type == "Integration Test":
        type_suffix = "IT"
    elif test_type == "API Test":
        type_suffix = "API"
    else:
        type_suffix = "UT"
    
    # Xử lý các trường hợp đặc biệt cho khớp với Spec cũ
    if test_type == "Integration Test":
        if feature == "Medicine" and layer == "Application":
            return "TC-CAT-MEDAPI-IT"
        elif feature == "Medicine":
            return "TC-CAT-MED-IT"
        elif feature == "Supplier":
            return "TC-PART-SUP-IT"
        elif feature == "Customer":
            return "TC-PART-CUS-IT"
    elif test_type == "API Test":
        if feature == "Medicine" or feature == "MedicineAppService":
            return "TC-CAT-MEDAPI-IT"
            
    # Unit Tests
    if feature == "Medicine":
        if "Manager" in class_name or layer == "DomainService":
            return "TC-CAT-MEDMGR-UT"
        return "TC-CAT-MED-UT"
    elif feature == "Supplier":
        if "Manager" in class_name:
            return "TC-PART-SUPMGR-UT"
        return "TC-PART-SUP-UT"
    elif feature == "Customer":
        if "Manager" in class_name:
            return "TC-PART-CUSMGR-UT"
        return "TC-PART-CUS-UT"
    elif feature == "SupplierProduct":
        return "TC-PART-SP-UT"
    elif feature == "SupplierProductCondition":
        return "TC-PART-SPC-UT"
        
    # Mặc định tự sinh
    feat_abbrev = feature_map.get(feature, feature[:3].upper())
    return f"TC-{context_prefix}-{feat_abbrev}-{type_suffix}"

def allocate_testcase_ids(test_cases):
    """
    Cấp phát ID tự động và lưu vết vào file mapping JSON để đảm bảo tính cố định.
    """
    # Nạp mapping hiện tại
    mapping = {}
    if os.path.exists(MAPPING_FILE):
        try:
            with open(MAPPING_FILE, "r", encoding="utf-8") as f:
                mapping = json.load(f)
        except Exception as e:
            print(f"WARNING: Could not parse mapping JSON ({e}). Initializing empty mapping.")
            
    updated = False
    
    for tc in test_cases:
        group_prefix = get_group_prefix(tc)
        method_name = tc["method_name"]
        
        # Đảm bảo nhóm tồn tại trong mapping
        if group_prefix not in mapping:
            mapping[group_prefix] = {}
            
        # Kiểm tra xem phương thức đã có ID chưa
        if method_name in mapping[group_prefix]:
            tc["testcase_id"] = mapping[group_prefix][method_name]
        else:
            # Sinh ID mới dựa trên số lượng hiện tại của nhóm
            num_existing = len(mapping[group_prefix])
            new_id = f"{group_prefix}-{num_existing + 1:03d}"
            mapping[group_prefix][method_name] = new_id
            tc["testcase_id"] = new_id
            updated = True
            print(f"Allocated New ID: {method_name} -> {new_id}")
            
    # Lưu lại file mapping nếu có cập nhật mới
    if updated or not os.path.exists(MAPPING_FILE):
        os.makedirs(os.path.dirname(MAPPING_FILE), exist_ok=True)
        with open(MAPPING_FILE, "w", encoding="utf-8") as f:
            json.dump(mapping, f, indent=2, ensure_ascii=False)
        print(f"SUCCESS: Saved updated testcase mapping to: {MAPPING_FILE}")
        
    return test_cases

def parse_trx_results():
    """
    Quét đệ quy và parse tất cả các file *.trx trong thư mục tập trung TRX_DIR để tổng hợp kết quả.
    """
    results = {}
    trx_files = []
    
    if not os.path.exists(TRX_DIR):
        print(f"WARNING: Centralized TRX directory not found: {TRX_DIR}")
        return results
        
    # Tìm kiếm các tệp .trx trong thư mục tập trung
    for root_path, _, files in os.walk(TRX_DIR):
        for file in files:
            if file.endswith(".trx"):
                trx_files.append(os.path.join(root_path, file))
                
    if not trx_files:
        print(f"WARNING: No .trx files found in: {TRX_DIR}")
        return results
        
    for trx_path in trx_files:
        try:
            tree = ET.parse(trx_path)
            root = tree.getroot()
            ns = {'': 'http://microsoft.com/schemas/VisualStudio/TeamTest/2010'}
            
            count = 0
            for result in root.findall('.//UnitTestResult', ns):
                test_name = result.get('testName') # Tên phương thức C#
                outcome = result.get('outcome')   # Passed, Failed, NotExecuted
                end_time_raw = result.get('endTime')  # Thời gian kết thúc
                
                # Format ngày chạy test
                test_date = datetime.now().strftime("%Y-%m-%d")
                if end_time_raw:
                    try:
                        # Lấy phần ngày YYYY-MM-DD
                        test_date = end_time_raw.split("T")[0]
                    except Exception:
                        pass
                        
                results[test_name] = {
                    "outcome": outcome,
                    "date": test_date
                }
                count += 1
            print(f"SUCCESS: Parsed {count} test execution records from: {trx_path}")
        except Exception as e:
            print(f"ERROR: Could not parse TRX file {trx_path} ({e}).")
            
    return results

def build_excel_report(test_cases, trx_results):
    """
    Kết xuất danh sách test cases ra Excel định dạng chuẩn Sleek Navy & Mint.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Scenarios & Cases"
    ws.views.sheetView[0].showGridLines = True
    
    # 1. Palette màu sắc thiết kế
    navy_header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    alt_row_fill = PatternFill(start_color="F2F4F7", end_color="F2F4F7", fill_type="solid")
    white_row_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    category_header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
    
    passed_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    passed_font = Font(name="Segoe UI", size=10, bold=True, color="375623")
    
    failed_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    failed_font = Font(name="Segoe UI", size=10, bold=True, color="C00000")
    
    pending_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    pending_font = Font(name="Segoe UI", size=10, bold=True, color="7F6000")
    
    header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    category_font = Font(name="Segoe UI", size=11, bold=True, color="000000")
    data_font = Font(name="Segoe UI", size=10, color="000000")
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # 2. Khởi tạo Header Columns
    headers = [
        "STT", "TestcaseID", "Bounded Context", "Feature/Component", "Layer", "Test Type", 
        "Priority", "Scenario Description", "Testcase Method Name", "Pre-conditions", 
        "TestData", "Expected Result", "Actual Result", "Status", "TestDate", "Note"
    ]
    
    ws.row_dimensions[1].height = 28
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = navy_header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        
    # 3. Phân nhóm và Sắp xếp Test Cases
    def get_sort_key(tc):
        is_partner = "Partner" in tc["filepath"] or tc["feature"] in ["Supplier", "Customer", "SupplierProduct", "SupplierProductCondition"]
        context = "Partner" if is_partner else "Catalog"
        return (context, tc["feature"], tc["layer"], tc["type"], tc["testcase_id"])
        
    sorted_cases = sorted(test_cases, key=get_sort_key)
    
    # 4. Ghi dữ liệu vào sheet
    current_row = 2
    last_feature = None
    stt_counter = 1
    
    for tc in sorted_cases:
        method_name = tc["method_name"]
        is_partner = "Partner" in tc["filepath"] or tc["feature"] in ["Supplier", "Customer", "SupplierProduct", "SupplierProductCondition"]
        context = "Partner" if is_partner else "Catalog"
        
        # Lấy thông tin chạy test thực tế từ TRX
        outcome = "Passed" # Mặc định
        test_date = datetime.now().strftime("%Y-%m-%d")
        
        if method_name in trx_results:
            trx_outcome = trx_results[method_name]["outcome"]
            test_date = trx_results[method_name]["date"]
            if trx_outcome == "Passed":
                outcome = "Passed"
            elif trx_outcome == "Failed":
                outcome = "Failed"
            else:
                outcome = "Pending"
        else:
            # Nếu không tìm thấy kết quả TRX -> kịch bản mới chưa được chạy
            outcome = "Pending"
            
        # Thêm Category Header phân cách trực quan
        current_feature = f"{context} - {tc['feature']}"
        category_header_text = f"▶ {current_feature.upper()} ({tc['layer'].upper()} LAYER - {tc['type'].upper()})"
        
        if category_header_text != last_feature:
            ws.row_dimensions[current_row].height = 24
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=16)
            
            cat_cell = ws.cell(row=current_row, column=1, value=category_header_text)
            cat_cell.fill = category_header_fill
            cat_cell.font = category_font
            cat_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            
            # Kẻ viền cho dòng gộp ô
            for col in range(1, 17):
                ws.cell(row=current_row, column=col).border = thin_border
                
            current_row += 1
            last_feature = category_header_text
            
        # Điền các giá trị đặc tả C#
        ws.row_dimensions[current_row].height = 20
        row_fill = alt_row_fill if stt_counter % 2 == 1 else white_row_fill
        
        preconditions = "Hệ thống ABP Framework đã khởi tạo đầy đủ."
        test_data = "Các khóa ngoại (GUID) và tham số nghiệp vụ hợp lệ."
        
        if "Duplicate" in method_name or "Exists" in method_name:
            preconditions = "Dữ liệu đã có sẵn bản ghi trùng lặp trong cơ sở dữ liệu."
            test_data = "Mã trùng hoặc Số đăng ký trùng."
        elif "NotFound" in method_name or "Invalid" in method_name:
            preconditions = "Khóa ngoại hoặc đối tượng liên kết không tồn tại."
            test_data = "Id ngẫu nhiên (Guid.NewGuid())."
            
        expected_result = f"Thực hiện thành công kịch bản: {tc['scenario']}"
        if "Throw" in method_name or "Exception" in method_name:
            expected_result = "Hệ thống ngăn chặn và ném ngoại lệ BusinessException mã lỗi chuẩn xác."
            
        actual_result = expected_result if outcome == "Passed" else "Ca kiểm thử thực thi thất bại. Vui lòng xem log chi tiết."
        
        row_values = [
            stt_counter,
            tc["testcase_id"],
            context,
            tc["feature"],
            tc["layer"],
            tc["type"],
            tc["priority"],
            tc["scenario"],
            method_name,
            preconditions,
            test_data,
            expected_result,
            actual_result,
            outcome,
            test_date,
            tc["note"]
        ]
        
        for col_idx, val in enumerate(row_values, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=val)
            cell.fill = row_fill
            cell.font = data_font
            cell.border = thin_border
            
            # Canh lề
            if col_idx in [1, 2, 3, 4, 5, 6, 7, 14, 15]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                
            # Định dạng màu cho Status & Actual Result
            if col_idx in [13, 14]:
                if outcome == "Passed":
                    cell.fill = passed_fill
                    cell.font = passed_font
                elif outcome == "Failed":
                    cell.fill = failed_fill
                    cell.font = failed_font
                else:
                    cell.fill = pending_fill
                    cell.font = pending_font
                    
        current_row += 1
        stt_counter += 1
        
    # 5. Cấu hình độ rộng cột tối ưu
    max_widths = {
        1: 6,   # STT
        2: 20,  # TestcaseID
        3: 15,  # Bounded Context
        4: 25,  # Feature
        5: 12,  # Layer
        6: 16,  # Test Type
        7: 10,  # Priority
        8: 45,  # Scenario Description
        9: 45,  # Testcase Method Name
        10: 35, # Pre-conditions
        11: 30, # TestData
        12: 45, # Expected Result
        13: 45, # Actual Result
        14: 12, # Status
        15: 12, # TestDate
        16: 25  # Note
    }
    
    for col_idx, width in max_widths.items():
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = width
        
    # Lưu file Excel
    os.makedirs(os.path.dirname(EXCEL_OUTPUT), exist_ok=True)
    wb.save(EXCEL_OUTPUT)
    print(f"SUCCESS: Generated Excel testcase specification at: {EXCEL_OUTPUT}")
    print(f"Total synchronized test cases: {len(sorted_cases)}")

def main():
    print("======================================================================")
    print("  QA TESTCASE SYNCHRONIZATION ENGINE (C# ATTRIBUTE -> EXCEL SPEC)")
    print("======================================================================")
    
    # 1. Quét mã nguồn C# tìm các QATest attributes
    print("Scanning C# test files for [QATest] attributes...")
    test_cases = scan_cs_files()
    print(f"Found {len(test_cases)} annotated test cases in source code.")
    
    if not test_cases:
        print("ERROR: No test cases found with [QATest] attribute! Make sure you annotated C# tests.")
        return
        
    # 2. Cấp phát và đồng bộ ID thông qua mapping JSON
    print("Allocating stable testcase IDs...")
    test_cases = allocate_testcase_ids(test_cases)
    
    # 3. Parse kết quả chạy test từ các tệp TRX XML trong thư mục tập trung
    print("Parsing dotnet test TRX results from centralized directory...")
    trx_results = parse_trx_results()
    
    # 4. Ghi đè file Excel Spec
    print("Compiling and writing to Excel workbook...")
    build_excel_report(test_cases, trx_results)
    
    print("======================================================================")
    print("  SYNCHRONIZATION COMPLETED SUCCESSFULLY!")
    print("======================================================================")

if __name__ == "__main__":
    main()
